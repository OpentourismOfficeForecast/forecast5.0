"""
This scripts aims to clean the "Arbeitsplan" provided by the tourist office
Since the original excel containes some name of the employees, it can not be shared publicly
The result is a csv file, consisting of numbers of employees needed for every day in 2016...2018
"""


from openpyxl import load_workbook
from calendar import monthrange
from math import floor
from datetime import datetime
import csv


def excel_style(col):
    """
    Convert given row and column number to an Excel-style cell name.
    Source: https://stackoverflow.com/questions/19153462/get-excel-style-column-names-from-column-number
    """
    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    result = []
    while col:
        col, rem = divmod(col-1, 26)
        result[:0] = letters[rem]
    return ''.join(result)


def workdays(number: int):
    """
    Every shift gets a number between 1 and 9, which must be translated to working hours
    :param number:
    :return: work days
    """
    assert 1 <= number <= 18
    if number >= 9:
        # Job location outside of Lenzerheide
        return 0
    elif 1 <= number < 5:  # Employee works full day in Lenzerheide
        return 1
    elif 5 <= number < 7:  # Employee works half day
        return 0.5
    elif 7 <= number < 9:  # Employee works full day
        return 1
    else:
        raise AssertionError


result = []  # store key value entries of date and man-days


for year in ["2016", "2017", "2018"]:
    print(year)
    wb = load_workbook("/home/andre_eggli/Desktop/datacleaning/raw_input/"
                       + year + "/Arbeitsplan " + year + ".xlsx")
    print(wb.sheetnames)
    for monthnbr, sheetname in enumerate(['Januar', 'Februar', 'März', 'April',
                                          'Mai', 'Juni', 'Juli', 'August',
                                          'September', 'Oktober', 'November', 'Dezember']):
        sheet = wb[sheetname]
        assert sheet["A1"].value == sheetname
        assert sheet["B5"].value == 1  # Start of month

        # find number of employees in Lenzerheide for that month
        number_of_employees = 0
        while True:
            # employee names start at row 6. Iterate column 'A' until empty row is found
            if sheet["A" + str(6+number_of_employees)].value is not None:
                number_of_employees += 1
            else:
                break

        # Iterate each date, find mandays in for this day
        number_days_in_month = monthrange(int(year), monthnbr+1)[1]
        for date in range(1, number_days_in_month+1):

            excelcolheader: str = excel_style(date + 1)  # day '1' -> Col "B"
            assert excelcolheader != "A"  # Col 'A' contains the name of the employee

            cnt_man_days: float = 0  # keep track of man-days within a day
            for employee in range(number_of_employees):
                number = sheet[excelcolheader + str(6+employee)].value

                # case distinction based on those codes in the bottom legend...
                if number is None:  # Employee not activated
                    pass
                elif type(number) == int:
                    cnt_man_days += workdays(number)
                elif type(number) == str:
                    number: str
                    if number == "TdS":  # Tour de Ski
                        # not counted here
                        pass
                    elif number == "D":  # Deskline
                        # No idea what this means, lets count it half
                        cnt_man_days += 0.5
                    elif number.startswith("Ei"):  # 7er Dienst
                        # No idea what this means, lets count it half
                        cnt_man_days += 0.5
                    elif number == "R" or number == "S":  # "Reserve / Schulung"
                        # not counted here
                        pass
                    elif number.startswith("8.1") or number.startswith("8.2"):
                        cnt_man_days += 0.5
                    else:  # all not considered cases.....
                        pass
                        print(number)
                elif type(number) == float:
                    # floor codes like '8.1' to '8.0'
                    cnt_man_days += workdays(int(floor(number)))

            # save this day in 'results'
            result.append([datetime(int(year), monthnbr+1, date), cnt_man_days])

# write result to csv
with open("/home/andre_eggli/Desktop/datacleaning/out/man_months_cleaned.csv", "w") as csvfile:
    csvwriter = csv.writer(csvfile,  delimiter=',')
    for point in result:
        csvwriter.writerow([point[0], point[1]])
